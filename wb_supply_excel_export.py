
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter



def export_supply_plan_to_excel(
    plan_result: dict[str, Any],
    output_path: str | Path,
    planned_target_days: float | int | None = None,
    max_supply: int | float | None = None,
) -> Path:
    """
    Экспортирует результат plan_supply_from_wb_items(...) в Excel-файл с 3 листами.

    Параметры
    ---------
    plan_result:
        Словарь, который вернула функция plan_supply_from_wb_items(...)

    output_path:
        Куда сохранить xlsx

    planned_target_days:
        Исходный горизонт "в планах".
        Важно: исходный код функции НЕ возвращает этот параметр в result,
        поэтому для заполнения шапки его нужно передать отдельно.
        Если не передан, в шапке будет пусто.

    max_supply:
        Исходный лимит поставки.
        Важно: исходный код функции НЕ возвращает этот параметр в result,
        поэтому для заполнения шапки его нужно передать отдельно.
        Если не передан, в шапке будет пусто.

    Возвращает
    ----------
    Path
        Путь к сохранённому xlsx-файлу.
    """
    output_path = Path(output_path)

    best_warehouse = plan_result.get("best_warehouse")
    achieved_days = plan_result.get("target_days")
    warehouse_scores = plan_result.get("warehouse_scores", {}) or {}
    items = plan_result.get("items", []) or []

    best_warehouse_score = warehouse_scores.get(best_warehouse) if best_warehouse else None
    items_in_supply_positions = sum(1 for item in items if (item.get("shipment_qty") or 0) > 0)
    total_shipment_units = plan_result.get("total_shipment")
    if total_shipment_units is None:
        total_shipment_units = sum(float(item.get("shipment_qty") or 0) for item in items)

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    ws_network = wb.create_sheet("По сети")
    ws_best_wh = wb.create_sheet("Лучший склад")
    ws_scores = wb.create_sheet("Склады")

    # -----------------------------
    # Styles
    # -----------------------------
    fill_title = PatternFill("solid", fgColor="1F4E78")
    fill_header = PatternFill("solid", fgColor="D9EAF7")
    fill_meta = PatternFill("solid", fgColor="E2F0D9")

    font_title = Font(color="FFFFFF", bold=True, size=12)
    font_header = Font(bold=True)
    font_meta = Font(bold=True)
    font_regular = Font(size=10)

    thin_gray = Side(style="thin", color="D0D0D0")
    border_bottom = Border(bottom=thin_gray)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    def _auto_width(ws):
        max_widths: dict[int, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                value_len = len(str(cell.value))
                max_widths[cell.column] = max(max_widths.get(cell.column, 0), value_len)
        for col_idx, width in max_widths.items():
            adjusted = min(max(width + 2, 12), 38)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

    def _style_header_row(ws, row_idx: int, cols_count: int):
        for col in range(1, cols_count + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = center
            cell.border = border_bottom

    def _is_number_like(v: Any) -> bool:
        return isinstance(v, (int, float)) and not isinstance(v, bool)

    def _write_value(cell, value, alignment=right):
        cell.value = value
        cell.alignment = alignment
        cell.font = font_regular

    def _sum_orders_14d(offices: dict[str, dict[str, Any]]) -> float:
        return float(sum((office.get("orders_14d", 0) or 0) for office in offices.values()))

    def _safe_count_value(v: Any) -> float:
        return float(v or 0)

    # -----------------------------
    # Sheet 1 — По сети
    # -----------------------------
    ws_network.merge_cells("A1:J1")
    c = ws_network["A1"]
    c.value = "План поставки по сети"
    c.fill = fill_title
    c.font = font_title
    c.alignment = left

    meta_rows = [
        ("Количество товаров в поставке", total_shipment_units),
        ("Количество позиций в поставке", items_in_supply_positions),
        ("Горизонт в планах", planned_target_days),
        ("Горизонт достигли", achieved_days),
        ("Лучший склад", best_warehouse),
        ("Скор лучшего склада", best_warehouse_score),
        ("Лимит поставки", max_supply),
    ]

    start_meta_row = 3
    for i, (label, value) in enumerate(meta_rows, start=start_meta_row):
        label_cell = ws_network.cell(row=i, column=1)
        value_cell = ws_network.cell(row=i, column=2)

        label_cell.value = label
        label_cell.fill = fill_meta
        label_cell.font = font_meta
        label_cell.alignment = left
        label_cell.border = border_bottom

        value_cell.value = value
        value_cell.alignment = right if _is_number_like(value) else left
        value_cell.border = border_bottom

    network_headers = [
        "nmID",
        "vendor",
        "Остаток по сети",
        "В пути по сети",
        "Эффективный остаток по сети",
        "Продажи за 14 дней по сети",
        "Продажи в день",
        "На сколько дней хватит без рассматриваемой поставки",
        "Количество в поставке",
        "На сколько дней хватит после поставки",
    ]

    header_row_network = start_meta_row + len(meta_rows) + 2
    for col_idx, header in enumerate(network_headers, start=1):
        ws_network.cell(row=header_row_network, column=col_idx, value=header)

    _style_header_row(ws_network, header_row_network, len(network_headers))
    ws_network.freeze_panes = f"A{header_row_network + 1}"

    sorted_items = sorted(
        items,
        key=lambda x: (
            -(x.get("shipment_qty") or 0),
            -(x.get("network_daily_sales") or 0),
            str(x.get("vendor") or ""),
        ),
    )

    for row_idx, item in enumerate(sorted_items, start=header_row_network + 1):
        nm_id = item.get("nmID")
        vendor = item.get("vendor")
        network_stock = _safe_count_value(item.get("network_stock"))
        network_in_transit = _safe_count_value(item.get("network_in_transit"))
        effective_network_stock = _safe_count_value(item.get("effective_network_stock"))
        network_daily_sales = _safe_count_value(item.get("network_daily_sales"))
        shipment_qty = _safe_count_value(item.get("shipment_qty"))
        orders_14d_network = _sum_orders_14d(item.get("offices", {}) or {})

        row_values = [
            nm_id,
            vendor,
            network_stock,
            network_in_transit,
            effective_network_stock,
            orders_14d_network,
            network_daily_sales,
            None,  # formula
            shipment_qty,
            None,  # formula
        ]

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws_network.cell(row=row_idx, column=col_idx)
            _write_value(cell, value, alignment=left if col_idx in (1, 2) else right)

        ws_network.cell(row=row_idx, column=8).value = (
            f'=IF(G{row_idx}=0,IF(E{row_idx}>0,"∞",""),E{row_idx}/G{row_idx})'
        )
        ws_network.cell(row=row_idx, column=10).value = (
            f'=IF(G{row_idx}=0,IF(E{row_idx}+I{row_idx}>0,"∞",""),(E{row_idx}+I{row_idx})/G{row_idx})'
        )

    # number formats
    for row_idx in range(header_row_network + 1, ws_network.max_row + 1):
        for col_idx in (3, 4, 5, 6, 9):
            ws_network.cell(row=row_idx, column=col_idx).number_format = '#,##0'
        for col_idx in (7, 8, 10):
            ws_network.cell(row=row_idx, column=col_idx).number_format = '0.00'

    if ws_network.max_row >= header_row_network:
        ws_network.auto_filter.ref = f"A{header_row_network}:J{ws_network.max_row}"

    # -----------------------------
    # Sheet 2 — Лучший склад
    # -----------------------------
    best_headers = [
        "nmID",
        "vendor",
        "Остаток на лучшем складе",
        "В пути на лучший склад",
        "Эффективный остаток на лучшем складе",
        "Продажи за 14 дней на лучшем складе",
        "Продажи в день",
        "На сколько дней хватит без рассматриваемой поставки",
        "Количество в поставке",
        "На сколько дней хватит после поставки",
    ]

    for col_idx, header in enumerate(best_headers, start=1):
        ws_best_wh.cell(row=1, column=col_idx, value=header)

    _style_header_row(ws_best_wh, 1, len(best_headers))
    ws_best_wh.freeze_panes = "A2"

    if best_warehouse:
        for row_idx, item in enumerate(sorted_items, start=2):
            office_info = (item.get("offices", {}) or {}).get(best_warehouse, {}) or {}

            nm_id = item.get("nmID")
            vendor = item.get("vendor")
            stock = _safe_count_value(office_info.get("stock"))
            in_transit = _safe_count_value(office_info.get("in_transit_qty"))
            effective_stock = _safe_count_value(office_info.get("effective_stock"))
            orders_14d = _safe_count_value(office_info.get("orders_14d"))
            daily_sales = _safe_count_value(office_info.get("daily_sales"))
            shipment_qty = _safe_count_value(item.get("shipment_qty"))

            row_values = [
                nm_id,
                vendor,
                stock,
                in_transit,
                effective_stock,
                orders_14d,
                daily_sales,
                None,  # formula
                shipment_qty,
                None,  # formula
            ]

            for col_idx, value in enumerate(row_values, start=1):
                cell = ws_best_wh.cell(row=row_idx, column=col_idx)
                _write_value(cell, value, alignment=left if col_idx in (1, 2) else right)

            ws_best_wh.cell(row=row_idx, column=8).value = (
                f'=IF(G{row_idx}=0,IF(E{row_idx}>0,"∞",""),E{row_idx}/G{row_idx})'
            )
            ws_best_wh.cell(row=row_idx, column=10).value = (
                f'=IF(G{row_idx}=0,IF(E{row_idx}+I{row_idx}>0,"∞",""),(E{row_idx}+I{row_idx})/G{row_idx})'
            )

        for row_idx in range(2, ws_best_wh.max_row + 1):
            for col_idx in (3, 4, 5, 6, 9):
                ws_best_wh.cell(row=row_idx, column=col_idx).number_format = '#,##0'
            for col_idx in (7, 8, 10):
                ws_best_wh.cell(row=row_idx, column=col_idx).number_format = '0.00'

        ws_best_wh.auto_filter.ref = f"A1:J{ws_best_wh.max_row}"
    else:
        ws_best_wh["A2"] = "Лучший склад не определён"
        ws_best_wh["A2"].alignment = left

    # -----------------------------
    # Sheet 3 — Склады
    # -----------------------------
    score_headers = ["Название склада", "Score склада"]

    for col_idx, header in enumerate(score_headers, start=1):
        ws_scores.cell(row=1, column=col_idx, value=header)

    _style_header_row(ws_scores, 1, len(score_headers))
    ws_scores.freeze_panes = "A2"

    sorted_scores = sorted(warehouse_scores.items(), key=lambda x: x[1], reverse=True)

    for row_idx, (warehouse_name, score) in enumerate(sorted_scores, start=2):
        cell_name = ws_scores.cell(row=row_idx, column=1, value=warehouse_name)
        cell_score = ws_scores.cell(row=row_idx, column=2, value=score)

        cell_name.alignment = left
        cell_score.alignment = right
        cell_name.font = font_regular
        cell_score.font = font_regular

    for row_idx in range(2, ws_scores.max_row + 1):
        ws_scores.cell(row=row_idx, column=2).number_format = '0.00'

    if ws_scores.max_row >= 1:
        ws_scores.auto_filter.ref = f"A1:B{ws_scores.max_row}"

    # -----------------------------
    # Common polish
    # -----------------------------
    for ws in (ws_network, ws_best_wh, ws_scores):
        _auto_width(ws)
        ws.sheet_view.showGridLines = False

    wb.save(output_path)
    return output_path


__all__ = ["export_supply_plan_to_excel"]
