from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook


def export_supply_barcodes_to_excel(
    plan_result: dict[str, Any],
    output_path: str | Path,
    items: list[dict[str, Any]] | None = None,
    barcode_by_nmid: dict[Any, Any] | None = None,
    sheet_name: str = "Поставка",
) -> Path:
    """
    Экспортирует Excel-файл с двумя колонками:
    - barcode
    - количество в поставке

    Без шапки/мета-блока: только строка с названиями колонок и данные.

    Важно:
    result из plan_supply_from_wb_items(...) barcode не содержит,
    поэтому barcode нужно передать:
    - либо через barcode_by_nmid,
    - либо через исходные items, если в них barcode есть.
    """
    output_path = Path(output_path)
    barcode_by_nmid = dict(barcode_by_nmid or {})
    result_items = plan_result.get("items", []) or []

    def _extract_first_scalar(value: Any) -> Any:
        if isinstance(value, (list, tuple)):
            return value[0] if value else None
        return value

    def _extract_barcode_from_item(item: dict[str, Any]) -> Any:
        direct_candidates = [
            item.get("barcode"),
            item.get("Barcode"),
            item.get("barcodeValue"),
            item.get("sku"),
            item.get("skus"),
            item.get("barcodes"),
        ]
        for candidate in direct_candidates:
            candidate = _extract_first_scalar(candidate)
            if candidate not in (None, ""):
                return candidate

        data = item.get("data", {}) or {}
        data_candidates = [
            data.get("barcode"),
            data.get("Barcode"),
            data.get("barcodeValue"),
            data.get("sku"),
            data.get("skus"),
            data.get("barcodes"),
        ]
        for candidate in data_candidates:
            candidate = _extract_first_scalar(candidate)
            if candidate not in (None, ""):
                return candidate

        sizes = data.get("sizes") or item.get("sizes") or []
        if isinstance(sizes, list):
            for size in sizes:
                if not isinstance(size, dict):
                    continue
                for key in ("barcode", "Barcode", "sku", "skus", "barcodes"):
                    candidate = _extract_first_scalar(size.get(key))
                    if candidate not in (None, ""):
                        return candidate

        return None

    if items:
        for item in items:
            nm_id = item.get("nmID")
            if nm_id is None or nm_id in barcode_by_nmid:
                continue
            barcode = _extract_barcode_from_item(item)
            if barcode not in (None, ""):
                barcode_by_nmid[nm_id] = barcode

    rows: list[tuple[str, int]] = []
    missing_nmid: list[Any] = []

    sorted_result_items = sorted(
        result_items,
        key=lambda x: (
            -(x.get("shipment_qty") or 0),
            str(x.get("vendor") or ""),
        ),
    )

    for item in sorted_result_items:
        shipment_qty = int(round(float(item.get("shipment_qty") or 0)))
        if shipment_qty <= 0:
            continue

        nm_id = item.get("nmID")
        barcode = barcode_by_nmid.get(nm_id)
        if barcode in (None, ""):
            barcode = _extract_barcode_from_item(item)

        if barcode in (None, ""):
            missing_nmid.append(nm_id)
            continue

        rows.append((str(barcode), shipment_qty))

    if missing_nmid:
        raise ValueError(
            "Не удалось определить barcode для nmID: " + ", ".join(map(str, missing_nmid))
        )

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    ws["A1"] = "Баркод"
    ws["B1"] = "Количество"

    for row_idx, (barcode, shipment_qty) in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=barcode)
        ws.cell(row=row_idx, column=1).number_format = "@"
        ws.cell(row=row_idx, column=2, value=shipment_qty)
        ws.cell(row=row_idx, column=2).number_format = '#,##0'

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 24
    ws.auto_filter.ref = f"A1:B{max(ws.max_row, 1)}"

    wb.save(output_path)
    return output_path


__all__ = ["export_supply_barcodes_to_excel"]
