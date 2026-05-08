import math
from pathlib import Path
from typing import Any
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference


def export_supply_plan_to_excel_warehouse(
    result_warehouse: dict[str, Any],
    name_report: str,
    days_to_plan: int | float,
) -> str:
    """
    Генерирует Excel-отчёт по результату calc_supply_for_warehouse(...).

    :param result_warehouse: результат функции calc_supply_for_warehouse
    :param name_report: имя файла или путь, например "supply_plan_tula.xlsx"
    :param days_to_plan: горизонт планирования в днях
    :return: путь к созданному xlsx-файлу
    """

    output_path = _normalize_xlsx_path(name_report)

    wb = Workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet("Итог")
    ws_all = wb.create_sheet("Все товары")
    ws_ship = wb.create_sheet("К отгрузке")

    _write_summary_sheet(ws_summary, result_warehouse, days_to_plan)
    _write_items_sheet(
        ws=ws_all,
        items=result_warehouse.get("items", []),
        days_to_plan=days_to_plan,
        only_need_to_ship=False,
        table_name="AllProductsTable",
    )
    _write_items_sheet(
        ws=ws_ship,
        items=result_warehouse.get("items", []),
        days_to_plan=days_to_plan,
        only_need_to_ship=True,
        table_name="ShipProductsTable",
    )

    _add_chart_to_summary(ws_summary, ws_ship)

    wb.save(output_path)

    return str(output_path)


def _write_summary_sheet(
    ws,
    result_warehouse: dict[str, Any],
    days_to_plan: int | float,
) -> None:
    warehouse = result_warehouse.get("warehouse", {})
    params = result_warehouse.get("params", {})
    summary = result_warehouse.get("summary", {})

    ws["A1"] = "План поставки на склад"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A1:D1")

    rows = [
        ("Склад", warehouse.get("officeName")),
        ("officeID", warehouse.get("officeID")),
        ("Горизонт планирования, дней", days_to_plan),
        ("Период продаж, дней", params.get("salesPeriodDays")),
        ("Товары в пути считаются приехавшими", "Да" if params.get("inTransitIsCountedAsArrived") else "Нет"),
        ("", ""),
        ("Количество товарных позиций", summary.get("itemsCount")),
        ("Остаток на складе, шт", summary.get("totalStockCount")),
        ("В пути на склад, шт", summary.get("totalInTransitCount")),
        ("Эффективный остаток, шт", summary.get("totalEffectiveStockCount")),
        ("Продажи за период, шт", summary.get("totalOrdersCount")),
        ("Продаж в день", summary.get("totalDailySales")),
        ("Целевой остаток, шт", summary.get("totalTargetStockCount")),
        ("Нужно отгрузить, шт", summary.get("totalNeedToShipCount")),
        ("Дней хватит текущего остатка", summary.get("totalDaysWithCurrentStock")),
        ("Дней хватит с учетом пути", summary.get("totalDaysWithInTransit")),
        ("Дней хватит после поставки", summary.get("totalDaysAfterShipment")),
    ]

    start_row = 3

    for idx, row in enumerate(rows, start=start_row):
        ws.cell(idx, 1).value = row[0]
        ws.cell(idx, 2).value = row[1]

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    for row in ws.iter_rows(min_row=start_row, max_row=start_row + len(rows) - 1, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            cell.border = _thin_border()

    for cell in ws["A"]:
        cell.font = Font(bold=True)

    important_cells = ["B14"]
    for cell_addr in important_cells:
        ws[cell_addr].fill = PatternFill("solid", fgColor="F4B183")
        ws[cell_addr].font = Font(bold=True)

    ws.freeze_panes = "A3"


def _write_items_sheet(
    ws,
    items: list[dict[str, Any]],
    days_to_plan: int | float,
    only_need_to_ship: bool,
    table_name: str,
) -> None:
    headers = [
        "nmID",
        "Артикул",
        "Дата данных",
        "officeID",
        "Склад",
        "Регион",
        "Есть склад в данных товара",
        "Остаток на складе",
        "В пути",
        "Эффективный остаток",
        "Продажи за 14 дней",
        "Сумма заказов",
        "Период продаж, дней",
        "Продаж в день",
        "Выкупы, шт",
        "Сумма выкупов",
        "Процент выкупа",
        "Горизонт, дней",
        "Целевой остаток",
        "К отгрузке",
        "Дней хватит текущего остатка",
        "Дней хватит с учетом пути",
        "Дней хватит после отгрузки",
        "Дефицит до поставки",
        "Излишек до поставки",
        "Ключи товара в пути",
    ]

    ws.append(headers)

    filtered_items = []

    for item in items:
        need_to_ship = _get_nested(item, "planning", "needToShipCount", default=0)

        if only_need_to_ship and need_to_ship <= 0:
            continue

        filtered_items.append(item)

    for item in filtered_items:
        sales_period_days = _get_nested(item, "sales", "salesPeriodDays", default=14)

        row = [
            item.get("nmID"),
            item.get("vendor"),
            item.get("timestamp"),

            _get_nested(item, "warehouse", "officeID"),
            _get_nested(item, "warehouse", "officeName"),
            _get_nested(item, "warehouse", "regionName"),
            _get_nested(item, "warehouse", "foundInProductOffices"),

            _get_nested(item, "stock", "stockCount", default=0),
            _get_nested(item, "stock", "inTransitCount", default=0),
            None,

            _get_nested(item, "sales", "ordersCount", default=0),
            _get_nested(item, "sales", "ordersSum", default=0),
            sales_period_days,
            None,

            _get_nested(item, "sales", "buyoutCount", default=0),
            _get_nested(item, "sales", "buyoutSum", default=0),
            _get_nested(item, "sales", "buyoutPercent", default=0),

            days_to_plan,
            None,
            None,
            None,
            None,
            None,
            None,
            None,

            ", ".join(_get_nested(item, "stock", "matchedTransitVendorKeys", default=[])),
        ]

        ws.append(row)

    max_row = ws.max_row

    if max_row >= 2:
        for row_num in range(2, max_row + 1):
            ws[f"J{row_num}"] = f"=H{row_num}+I{row_num}"
            ws[f"N{row_num}"] = f'=IF(M{row_num}>0,K{row_num}/M{row_num},"")'
            ws[f"S{row_num}"] = f'=ROUNDUP(N{row_num}*R{row_num},0)'
            ws[f"T{row_num}"] = f"=MAX(0,S{row_num}-J{row_num})"
            ws[f"U{row_num}"] = f'=IF(N{row_num}>0,H{row_num}/N{row_num},"")'
            ws[f"V{row_num}"] = f'=IF(N{row_num}>0,J{row_num}/N{row_num},"")'
            ws[f"W{row_num}"] = f'=IF(N{row_num}>0,(J{row_num}+T{row_num})/N{row_num},"")'
            ws[f"X{row_num}"] = f"=MAX(0,S{row_num}-J{row_num})"
            ws[f"Y{row_num}"] = f"=MAX(0,J{row_num}-S{row_num})"

    _style_items_sheet(ws)

    if max_row >= 2:
        table_range = f"A1:{get_column_letter(ws.max_column)}{max_row}"
        table = Table(displayName=table_name, ref=table_range)

        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        table.tableStyleInfo = style
        ws.add_table(table)
    else:
        ws["A2"] = "Нет товаров для отображения"


def _style_items_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    border = _thin_border()

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = {
        "A": 13,  # nmID
        "B": 16,  # vendor
        "C": 20,  # timestamp
        "D": 12,  # officeID
        "E": 28,  # officeName
        "F": 28,  # regionName
        "G": 16,

        "H": 14,
        "I": 12,
        "J": 18,
        "K": 16,
        "L": 14,
        "M": 14,
        "N": 14,
        "O": 12,
        "P": 14,
        "Q": 14,
        "R": 14,
        "S": 16,
        "T": 14,
        "U": 18,
        "V": 18,
        "W": 18,
        "X": 16,
        "Y": 16,
        "Z": 28,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row_num in range(2, ws.max_row + 1):
        for col in ["H", "I", "J", "K", "L", "M", "O", "P", "R", "S", "T", "X", "Y"]:
            ws[f"{col}{row_num}"].number_format = "0"

        for col in ["N", "Q", "U", "V", "W"]:
            ws[f"{col}{row_num}"].number_format = "0.00"

    if ws.max_row >= 2:
        ws.conditional_formatting.add(
            f"T2:T{ws.max_row}",
            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                fill=PatternFill("solid", fgColor="F4B183"),
            ),
        )

        ws.conditional_formatting.add(
            f"V2:V{ws.max_row}",
            CellIsRule(
                operator="lessThan",
                formula=["$R2"],
                fill=PatternFill("solid", fgColor="F8CBAD"),
            ),
        )

    ws.sheet_view.showGridLines = False


def _add_chart_to_summary(ws_summary, ws_ship) -> None:
    """
    Добавляет простой график по товарам к отгрузке.
    Берёт первые 15 строк с листа 'К отгрузке'.
    """

    if ws_ship.max_row < 2:
        return

    max_chart_row = min(ws_ship.max_row, 16)

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Топ товаров к отгрузке"
    chart.y_axis.title = "Артикул"
    chart.x_axis.title = "Количество"

    data = Reference(ws_ship, min_col=20, min_row=1, max_row=max_chart_row)  # T: К отгрузке
    cats = Reference(ws_ship, min_col=2, min_row=2, max_row=max_chart_row)   # B: Артикул

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    chart.height = 8
    chart.width = 16

    ws_summary.add_chart(chart, "D3")


def _normalize_xlsx_path(name_report: str) -> Path:
    if not name_report:
        name_report = "supply_plan_warehouse.xlsx"

    path = Path(name_report)

    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")

    if path.parent and str(path.parent) != ".":
        os.makedirs(path.parent, exist_ok=True)

    return path


def _get_nested(data: dict[str, Any], *keys: str, default=None):
    current = data

    for key in keys:
        if not isinstance(current, dict):
            return default

        current = current.get(key)

        if current is None:
            return default

    return current


def _thin_border() -> Border:
    side = Side(style="thin", color="D9E2F3")

    return Border(
        left=side,
        right=side,
        top=side,
        bottom=side,
    )




















import os
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def export_supply_barcodes_to_excel_warehouse(
    result: dict[str, Any],
    name_export: str,
    barcode_by_nmid: dict[int | str, str | int | list[str] | tuple[str, ...]],
) -> str:
    """
    Генерирует Excel-файл для загрузки поставки на WB.

    Вид строго как в приложенном файле:

    Лист: Поставка

    Баркод | Количество
    1234567890123 | 10
    1234567890124 | 5

    :param result: результат функции calc_supply_for_warehouse(...)
    :param name_export: имя файла или путь, например "wb_supply_upload.xlsx"
    :param barcode_by_nmid: словарь {nmID: barcode}
                            ключ nmID может быть int или str
    :return: путь к созданному xlsx-файлу
    """

    output_path = _normalize_xlsx_path(name_export)

    wb = Workbook()
    ws = wb.active
    ws.title = "Поставка"

    ws["A1"] = "Баркод"
    ws["B1"] = "Количество"

    ws["A1"].font = Font(bold=False)
    ws["B1"].font = Font(bold=False)

    ws["A1"].alignment = Alignment(horizontal="left")
    ws["B1"].alignment = Alignment(horizontal="left")

    row_num = 2

    missing_barcodes = []
    skipped_zero_qty = []

    for item in result.get("items", []):
        nm_id = item.get("nmID")
        vendor = item.get("vendor")

        qty = _get_nested(
            item,
            "planning",
            "needToShipCount",
            default=0,
        )

        qty = _safe_int(qty)

        if qty <= 0:
            skipped_zero_qty.append(
                {
                    "nmID": nm_id,
                    "vendor": vendor,
                    "qty": qty,
                }
            )
            continue

        barcode = _find_barcode_by_nmid(nm_id, barcode_by_nmid)

        if not barcode:
            missing_barcodes.append(
                {
                    "nmID": nm_id,
                    "vendor": vendor,
                    "qty": qty,
                }
            )
            continue

        ws.cell(row=row_num, column=1).value = str(barcode)
        ws.cell(row=row_num, column=2).value = qty

        row_num += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12

    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")

    # Важно: WB часто нормально читает баркоды как число,
    # но чтобы не потерять точность и не получить научный формат,
    # баркод лучше хранить как текст.
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=1).number_format = "@"
        ws.cell(row=row, column=2).number_format = "0"

    # Если есть товары к отгрузке без barcode — не молчим.
    # Файл всё равно создаём, но добавляем технический лист с ошибками.
    if missing_barcodes:
        ws_errors = wb.create_sheet("Ошибки")

        ws_errors.append(["nmID", "Артикул", "Количество", "Ошибка"])

        for row in missing_barcodes:
            ws_errors.append(
                [
                    row["nmID"],
                    row["vendor"],
                    row["qty"],
                    "Не найден barcode в barcode_by_nmid",
                ]
            )

        ws_errors.column_dimensions["A"].width = 14
        ws_errors.column_dimensions["B"].width = 18
        ws_errors.column_dimensions["C"].width = 14
        ws_errors.column_dimensions["D"].width = 36

    wb.save(output_path)

    return str(output_path)

def _find_barcode_by_nmid(
    nm_id: int | str,
    barcode_by_nmid: dict[int | str, Any],
) -> str | None:
    """
    Ищет barcode по nmID.

    Поддерживает ключи:
    - int: 66837050
    - str: "66837050"

    Поддерживает значения:
    - "2023293531004"
    - 2023293531004
    - ["2023293531004", "другой"] — берём первый
    """

    if nm_id is None:
        return None

    possible_keys = [
        nm_id,
        str(nm_id),
    ]

    try:
        possible_keys.append(int(nm_id))
    except (TypeError, ValueError):
        pass

    barcode = None

    for key in possible_keys:
        if key in barcode_by_nmid:
            barcode = barcode_by_nmid[key]
            break

    if barcode is None:
        return None

    if isinstance(barcode, (list, tuple, set)):
        barcode_list = list(barcode)

        if not barcode_list:
            return None

        barcode = barcode_list[0]

    barcode = str(barcode).strip()

    if not barcode:
        return None

    return barcode


def _get_nested(data: dict[str, Any], *keys: str, default=None):
    current = data

    for key in keys:
        if not isinstance(current, dict):
            return default

        current = current.get(key)

        if current is None:
            return default

    return current


def _safe_int(value: Any) -> int:
    if value is None:
        return 0

    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return 0


def _normalize_xlsx_path(name: str) -> Path:
    if not name:
        name = "wb_supply_upload.xlsx"

    path = Path(name)

    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")

    if path.parent and str(path.parent) != ".":
        os.makedirs(path.parent, exist_ok=True)

    return path