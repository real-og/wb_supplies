import json
from datetime import datetime, date
from openpyxl import load_workbook


def generate_sales_json(
    excel_path: str,
    sheet_name: str | None = None,
    date_column_name: str = "День",
    sales_column_name: str = "Заказано всего, шт.",
    output_path: str | None = None,
) -> str:
    """
    Читает Excel-файл и генерирует JSON со структурой:
    {
        "YYYY-MM-DD": количество_продаж
    }

    :param excel_path: путь к xlsx-файлу
    :param sheet_name: имя листа; если None — берется первый лист
    :param date_column_name: название колонки с датой
    :param sales_column_name: название колонки с продажами
    :param output_path: если указан, JSON будет сохранен в файл
    :return: JSON-строка
    """

    wb = load_workbook(excel_path, data_only=True)

    if sheet_name is None:
        ws = wb[wb.sheetnames[0]]
    else:
        ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Лист пустой")

    headers = list(rows[0])

    try:
        date_idx = headers.index(date_column_name)
    except ValueError:
        raise ValueError(f"Не найдена колонка с датой: {date_column_name}")

    try:
        sales_idx = headers.index(sales_column_name)
    except ValueError:
        raise ValueError(f"Не найдена колонка с продажами: {sales_column_name}")

    result = {}

    for row in rows[1:]:
        if row is None:
            continue

        day_value = row[date_idx] if date_idx < len(row) else None
        sales_value = row[sales_idx] if sales_idx < len(row) else None

        if day_value is None or sales_value is None:
            continue

        # Приводим дату к строке YYYY-MM-DD
        if isinstance(day_value, datetime):
            day_str = day_value.strftime("%Y-%m-%d")
        elif isinstance(day_value, date):
            day_str = day_value.strftime("%Y-%m-%d")
        else:
            # если в ячейке уже строка
            day_str = str(day_value).strip()

        # Приводим продажи к int
        try:
            sales_count = int(sales_value)
        except (TypeError, ValueError):
            continue

        result[day_str] = sales_count

    json_result = json.dumps(result, ensure_ascii=False, indent=4)

    if output_path:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(json_result)

    return json_result

if __name__ == "__main__":
    generate_sales_json('report_coof.xlsx', output_path='report_coof.json')